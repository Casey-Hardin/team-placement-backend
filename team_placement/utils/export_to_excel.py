# native import
from io import BytesIO

# third-party imports
from fastapi import Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter

# external imports
from team_placement.schemas import Cell


def export_to_excel(cells: list[list[Cell]]) -> None:
    """Sends raw data to an Excel file."""
    # create an object in memory for the workbook
    file = BytesIO()

    # create a new workbook
    wb = Workbook()

    # collect the first worksheet
    ws = wb.active

    # border styles
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    blueFill = PatternFill(start_color="3ea6eb", end_color="3ea6eb", fill_type="solid")

    # output data to Excel
    print("Creating an Excel File.")
    for row_index, row in enumerate(cells, 1):
        column = 1
        for cell_dict in row:
            try:
                value = float(cell_dict.value)
            except:
                value = cell_dict.value
            cell = ws.cell(row=row_index, column=column, value=value)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.border = thin_border

            # color merged header cells
            if cell_dict.colspan != 1 and cell_dict.value != "":
                cell.fill = blueFill

            # merge cells
            if cell_dict.colspan != 1:
                ws.merge_cells(
                    start_row=row_index,
                    start_column=column,
                    end_row=row_index,
                    end_column=column + cell_dict.colspan - 1,
                )
            column = column + cell_dict.colspan

    # size columns
    for index in range(1, column + 1):
        letter = get_column_letter(index)
        ws.column_dimensions[letter].width = 20
    wb.save(file)
    wb.close()
    print("Excel file created successfully.")

    # return Excel file content for download
    file.seek(0)
    return Response(content=file.read())
