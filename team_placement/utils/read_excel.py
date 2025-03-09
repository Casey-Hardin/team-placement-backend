# native imports
from datetime import datetime
from enum import Enum
from io import BytesIO

# third-party imports
from dateutil.relativedelta import relativedelta
from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook
import shortuuid

# external imports
from team_placement.constants import FIRST_TIME_COST
from team_placement.schemas import BooleanEnum, Collective, Gender, Person


class Columns(str, Enum):
    first_name = "first_name"
    last_name = "last_name"
    gender = "gender"
    preferred = "preferred"
    paid = "paid"
    donation = "donation"
    birthday = "birthday"
    collective = "collective"
    leader = "leader"
    participant = "participant"


# column identifier : column name
COLUMNS_DICT = {
    Columns.first_name: "First Name",
    Columns.last_name: "Last Name",
    Columns.gender: "Gender",
    Columns.preferred: (
        "At the retreat, we have to determine which room each person is staying in. "
        "Is there anyone that you want to room with this weekend?"
    ),
    Columns.paid: "Form Total",
    Columns.donation: "Donation",
    Columns.birthday: "Birthday",
    Columns.collective: "How many times have you attended a Collective on Thursday Night?",
}

# allow leaders to be set through the interface rather than required in the input file
OPTIONAL_COLUMNS_DICT = {
    Columns.leader: "Team Name",
    Columns.participant: "Participating on a Team",
}


def read_excel(file: UploadFile) -> tuple[list[Person], str]:
    """
    Reads people from an Excel file.

    Parameters
    ----------
    file
        Excel file with people data.

    Returns
    -------
    list[Person]
        List of people collected from the Excel file.
    str
        Message indicating any issues with the input file.
    """
    # file must have a valid extension
    if not any([file.filename.endswith(x) for x in [".xlsx", ".csv"]]):
        message = "Input must be an Excel file!"
        print(message)
        raise HTTPException(status_code=412, detail={"message": message})

    # read Excel file
    contents = BytesIO(file.file.read())
    workbook = load_workbook(filename=contents, read_only=True)
    sheet = workbook.active

    # find columns from row 5 of the active worksheet - the header row
    indices = {}
    missing_columns = []
    for code, name in (COLUMNS_DICT | OPTIONAL_COLUMNS_DICT).items():
        # find the first column with the column name
        column_index = next(
            (i for i, x in enumerate([y.value for y in sheet[5]]) if x == name), None
        )

        # collect names of missing required columns
        if column_index is None:
            if name in OPTIONAL_COLUMNS_DICT.values():
                continue
            missing_columns.append(name)
            continue
        indices[code] = column_index

    # required columns must be found
    if missing_columns != []:
        message = (
            "Columns\n"
            + "\n".join(missing_columns)
            + "\nare missing from row 5 of the input file!"
        )
        print(message)
        raise HTTPException(status_code=413, detail={"message": message})

    people, message = [], ""

    # process data values
    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        # row 5 is the header row
        if row_index < 6:
            continue

        # skip empty rows
        if all([x in ["", None] for x in row]):
            continue

        # collect first and last names
        first_name = row[indices[Columns.first_name]]
        last_name = row[indices[Columns.last_name]]
        if first_name in ["", None] or last_name in ["", None]:
            message += (
                f"First and / or Last Name is missing for person at row {row_index}.\n"
            )
            continue
        first_name = str(first_name).strip().capitalize()
        last_name = str(last_name).strip().capitalize()

        # collect gender
        gender: str = row[indices[Columns.gender]]
        if gender.lower().startswith("m") or gender.lower() == "dude":
            gender = Gender.male.value
        elif gender.lower().startswith("f") or gender.lower() in ["girl", "lady"]:
            gender = Gender.female.value
        if gender not in [x.value for x in Gender]:
            message += (
                f"Gender is missing for {first_name}, {last_name} at row {row_index}.\n"
            )
            continue

        # collect a string to decipher preferred
        # people once all people are read
        preferred_people_raw = (
            row[indices[Columns.preferred]]
            if row[indices[Columns.preferred]] is not None
            else ""
        )

        # collect first-time status based on amount paid
        paid = (
            str(row[indices[Columns.paid]]).replace("$", "").replace(",", "")
            if row[indices[Columns.paid]] is not None
            else 0
        )
        donation = (
            str(row[indices[Columns.donation]]).replace("$", "").replace(",", "")
            if row[indices[Columns.donation]] is not None
            else 0
        )
        try:
            cost = float(paid) - float(donation) if paid != "-" else 0
            first_time = BooleanEnum.yes if cost <= FIRST_TIME_COST else BooleanEnum.no
        except:
            message += f"Paid and / or Donation are missing for {first_name}, {last_name} at row {row_index}.\n"
            continue

        # collect age based on birthday
        birthday = row[indices[Columns.birthday]]
        try:
            age = relativedelta(datetime.today(), birthday).years
        except:
            message += (
                f"Age is missing for {first_name}, {last_name} at row {row_index}.\n"
            )
            continue

        # collect collective status
        collective = row[indices[Columns.collective]].split(",")[0]
        if collective not in [x.value for x in Collective]:
            message += f"Collective Status is missing for {first_name}, {last_name} at row {row_index}.\n"
            continue

        # collect leader status
        try:
            team = (
                row[indices[Columns.leader]]
                if row[indices[Columns.leader]] is not None
                else ""
            )
        except:
            team = ""

        leader = BooleanEnum.yes if team != "" else BooleanEnum.no

        # collect participant status
        try:
            participant = row[indices[Columns.participant]]
            participant = (
                participant if participant == BooleanEnum.no else BooleanEnum.yes
            )
        except:
            participant = BooleanEnum.yes

        # create a person
        person = Person(
            index=shortuuid.ShortUUID().random(length=10),
            order=row_index - 5,
            firstName=first_name,
            lastName=last_name,
            age=age,
            gender=gender,
            firstTime=first_time,
            collective=collective,
            preferredPeopleRaw=preferred_people_raw,
            preferredPeople=[],
            leader=leader,
            team=team,
            room="",
            participant=participant,
        )
        people.append(person)
    return people, message
