# native imports
from collections import defaultdict
from datetime import datetime
from io import BytesIO

# third-party imports
from dateutil.relativedelta import relativedelta
from fastapi import UploadFile
from openpyxl import load_workbook

# external imports
from .objects import Person
from . import schemas


def read_excel(file: UploadFile) -> tuple[list[Person], dict[int: list[Person]], str]:
    people, leaders, message = [], defaultdict(lambda: []), ""

    # read the XLSX file
    contents = BytesIO(file.file.read())
    workbook = load_workbook(filename=contents, read_only=True)
    sheet = workbook.active

    # column identifier : column name
    columns = {
        "first_name": "First Name",
        "last_name": "Last Name",
        "gender": "Gender",
        "preferred": (
            "At the retreat, we have to determine which room each person is staying in. "
            "Is there anyone that you want to room with this weekend?"
        ),
        "paid": "Form Total",
        "donation": "Donation",
        "birthday": "Birthday",
        "collective": "How many times have you attended a Collective on Thursday Night?",
    }

    # allow leaders to be set through the interface rather than required in the input file
    optional_columns = {
        "leader": "Leader Team"
    }

    # find columns from row 5 of the active worksheet - the header row
    indices = {}
    missing_columns = []
    for code, name in (columns | optional_columns).items():
        column_index = next((i for i, x in enumerate([y.value for y in sheet[5]]) if x == name), None)
        if column_index is None:
            if name in optional_columns:
                continue
            missing_columns.append(name)
            continue
        indices[code] = column_index

    # required columns must be found
    if missing_columns != []:
        message += (
            "Columns\n"
            + "\n".join(missing_columns)
            + "\nare missing from row 5 of the input file!"
        )
        return people, message

    # process data values
    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        # row 5 is the header row
        if row_index < 6:
            continue

        # collect first and last names
        first_name = row[indices["first_name"]]
        last_name = row[indices["last_name"]]
        if first_name in ["", None] or last_name in ["", None]:
            continue

        # collect gender
        gender = row[indices["gender"]]
        if gender not in [x.value for x in schemas.Gender]:
            message += f"Gender is missing for {first_name}, {last_name}.\n"
            continue

        # collect a string to decipher preferred
        # people once all people are read
        preferred_people_raw = row[indices["preferred"]]

        # collect first-time status based on amount paid
        paid = str(row[indices["paid"]]).replace("$", "").replace(",", "") if row[indices["paid"]] is not None else 0
        donation = str(row[indices["donation"]]).replace("$", "").replace(",", "") if row[indices["donation"]] is not None else 0
        try:
            cost = float(paid) - float(donation)
            first_time = True if cost <= 35 else False
        except:
            message += f"Paid and / or Donation are missing for {first_name}, {last_name}.\n"
            continue

        # collect age based on birthday
        birthday = row[indices["birthday"]]
        try:
            age = relativedelta(datetime.today(), birthday).years
        except:
            message += f"Age is missing for {first_name}, {last_name}.\n"
            continue

        # collect collective status
        collective = row[indices["collective"]]
        if collective not in [x.value for x in schemas.Collective]:
            message += f"Collective Status is missing for {first_name}, {last_name}.\n"
            continue

        # collect leader status
        try:
            team_number = int(row[indices["leader"]])
        except:
            team_number = None

        # create a person
        person = Person(schemas.Person(**{
            "first_name": first_name,
            "last_name": last_name,
            "gender": gender,
            "preferred_people_raw": preferred_people_raw,
            "first_time": first_time,
            "age": age,
            "collective": collective,
        }))
        people.append(person)
        if team_number is not None:
            leaders[team_number].append(person)

    # decipher preferred people
    for person in people:
        # container to append preferred people
        preferred = []

        # collect the raw string
        # all comparisons are case-insensitive
        text = person.preferred_people_raw
        if text is None:
            continue
        text = text.lower()

        # remove special characters
        omit = {
            "!",
            "?",
        }
        for word in omit:
            text = text.replace(word, "")

        # replace path separators
        text = text.replace(" and ", ", ").replace(" or ", ", ").replace("/", ", ")

        # interpret names from raw strings
        search_array = [x.strip() for x in text.split(",") if x != ""]
        for full_name in search_array:
            # first and last name found
            if " " in full_name and len(full_name.split(" ")) == 2:
                first_name, last_name = full_name.split(" ")

                # match by strictly first and last name
                matches = [
                    x for x in people
                    if first_name == x.first_name.lower()
                    and last_name == x.last_name.lower()
                ]

                if matches == []:
                    # match by first name or nickname
                    # and by first initial of last name
                    matches = [
                        x for x in people
                        if (
                            first_name == x.first_name.lower()
                            or (
                                x.nicknames is not None
                                and any([
                                    first_name == name.lower()
                                    for name in x.nicknames
                                ])
                            )
                        )
                        and x.last_name.lower().startswith(last_name[0])
                    ]
            else:
                # no last names found - build list of names
                names = [full_name]
                if " " in full_name:
                    names = [x for x in full_name.split(" ") if x != ""]

                # match based on first name or nickname
                # a match must also either have the same last name,
                # have also picked this person
                # or be the only individual with that first name or nickname
                matches = []
                for name in names:
                    matches += [
                        x for x in people
                        if (
                            name == x.first_name.lower()
                            or (
                                x.nicknames is not None
                                and any([
                                    name == nickname.lower()
                                    for nickname in x.nicknames
                                ])
                            )
                        ) and (
                            person.last_name == x.last_name
                            or (
                                x.preferred_people_raw != None
                                and (
                                    person.first_name.lower() in x.preferred_people_raw.lower()
                                    or (
                                        person.nicknames is not None
                                        and any([
                                            nickname.lower() in x.preferred_people_raw.lower()
                                            for nickname in person.nicknames
                                        ])
                                    )
                                )
                            )
                            or len([
                                x for x in people
                                if name == x.first_name.lower()
                            ]) == 1
                            or len([
                                x for x in people if x.nicknames is not None
                                and any([
                                    name == nickname.lower()
                                    for nickname in x.nicknames
                                ])
                            ]) == 1
                        )
                    ]

            # add unique matches to a person's preferred list
            # a person cannot pick themselves
            for match in matches:
                if match not in preferred and match != person:
                    preferred.append(match)

        # assign preferred people
        person.preferred_people = preferred
    return people, leaders, message